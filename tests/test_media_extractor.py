"""测试媒体提取器"""

from io import BytesIO
from pathlib import Path
from unittest.mock import Mock, patch

import openpyxl
import pytest
from PIL import Image

from table_parsing.ir.model import MediaObject
from table_parsing.media.extractor import MediaExtractor


@pytest.fixture
def sample_image():
    """创建示例图片"""
    img = Image.new("RGB", (100, 100), color="red")
    img_io = BytesIO()
    img.save(img_io, format="PNG")
    return img_io.getvalue()


@pytest.fixture
def xlsx_with_image(tmp_path: Path, sample_image: bytes):
    """创建包含图片的 XLSX 文件"""
    file_path = tmp_path / "with_image.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # 添加一些文本
    ws["A1"] = "标题"
    ws["A2"] = "数据"

    # 添加图片到工作表
    img = openpyxl.drawing.image.Image(BytesIO(sample_image))
    img.anchor = "C3"
    ws.add_image(img)

    wb.save(file_path)
    return file_path


@pytest.fixture
def xlsx_with_multiple_images(tmp_path: Path, sample_image: bytes):
    """创建包含多张图片的 XLSX 文件"""
    file_path = tmp_path / "multiple_images.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "多图测试"

    ws["A1"] = "Sheet with multiple images"

    # 添加第一张图片
    img1 = openpyxl.drawing.image.Image(BytesIO(sample_image))
    img1.anchor = "A3"
    ws.add_image(img1)

    # 添加第二张图片
    img2 = openpyxl.drawing.image.Image(BytesIO(sample_image))
    img2.anchor = "E5"
    ws.add_image(img2)

    wb.save(file_path)
    return file_path


@pytest.fixture
def xlsx_no_images(tmp_path: Path):
    """创建没有图片的 XLSX 文件"""
    file_path = tmp_path / "no_images.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "纯文本"

    ws["A1"] = "只有文本"
    ws["A2"] = "没有图片"

    wb.save(file_path)
    return file_path


class TestMediaExtractor:
    """测试 MediaExtractor 基础功能"""

    def test_extractor_creation(self):
        """测试创建提取器实例"""
        extractor = MediaExtractor()
        assert extractor is not None

    def test_extract_from_xlsx_with_images(self, xlsx_with_image: Path):
        """测试从包含图片的 XLSX 文件提取"""
        extractor = MediaExtractor()
        media_objects = extractor.extract_from_xlsx(xlsx_with_image)

        assert len(media_objects) == 1
        assert isinstance(media_objects[0], MediaObject)
        assert media_objects[0].type == "image"
        assert media_objects[0].anchor_row == 3
        assert media_objects[0].anchor_col == 3
        assert media_objects[0].raw_data is not None
        assert media_objects[0].raw_format is not None

    def test_extract_multiple_images(self, xlsx_with_multiple_images: Path):
        """测试提取多张图片"""
        extractor = MediaExtractor()
        media_objects = extractor.extract_from_xlsx(xlsx_with_multiple_images)

        assert len(media_objects) == 2

        # 验证第一张图片
        assert media_objects[0].type == "image"
        assert media_objects[0].anchor_row == 3
        assert media_objects[0].anchor_col == 1

        # 验证第二张图片
        assert media_objects[1].type == "image"
        assert media_objects[1].anchor_row == 5
        assert media_objects[1].anchor_col == 5

    def test_extract_from_xlsx_without_images(self, xlsx_no_images: Path):
        """测试从没有图片的 XLSX 文件提取"""
        extractor = MediaExtractor()
        media_objects = extractor.extract_from_xlsx(xlsx_no_images)

        assert len(media_objects) == 0

    def test_media_object_has_raw_data(self, xlsx_with_image: Path):
        """测试 MediaObject 包含原始图片数据"""
        extractor = MediaExtractor()
        media_objects = extractor.extract_from_xlsx(xlsx_with_image)

        assert len(media_objects) == 1
        assert media_objects[0].raw_data is not None
        assert isinstance(media_objects[0].raw_data, bytes)
        assert len(media_objects[0].raw_data) > 0

    def test_media_object_has_format(self, xlsx_with_image: Path):
        """测试 MediaObject 包含图片格式"""
        extractor = MediaExtractor()
        media_objects = extractor.extract_from_xlsx(xlsx_with_image)

        assert len(media_objects) == 1
        assert media_objects[0].raw_format is not None
        assert isinstance(media_objects[0].raw_format, str)
        # PNG 格式
        assert media_objects[0].raw_format.lower() in ("png", "image/png")

    def test_anchor_position_detection(self, xlsx_with_multiple_images: Path):
        """测试锚点位置检测"""
        extractor = MediaExtractor()
        media_objects = extractor.extract_from_xlsx(xlsx_with_multiple_images)

        # 第一张图片在 A3 (row=3, col=1)
        assert media_objects[0].anchor_row == 3
        assert media_objects[0].anchor_col == 1

        # 第二张图片在 E5 (row=5, col=5)
        assert media_objects[1].anchor_row == 5
        assert media_objects[1].anchor_col == 5

    def test_extract_handles_invalid_file(self, tmp_path: Path):
        """测试处理无效文件"""
        invalid_file = tmp_path / "invalid.xlsx"
        invalid_file.write_text("Not a valid XLSX")

        extractor = MediaExtractor()

        # 应该抛出异常或返回空列表
        # 这里我们期望抛出异常
        with pytest.raises(Exception):
            extractor.extract_from_xlsx(invalid_file)


class TestMediaExtractorImageFormats:
    """测试不同图片格式的提取"""

    def test_extract_jpeg_image(self, tmp_path: Path):
        """测试提取 JPEG 图片"""
        # 创建 JPEG 图片
        img = Image.new("RGB", (50, 50), color="blue")
        img_io = BytesIO()
        img.save(img_io, format="JPEG")
        jpeg_bytes = img_io.getvalue()

        # 创建 XLSX 文件
        file_path = tmp_path / "jpeg_image.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "JPEG Image"

        img_obj = openpyxl.drawing.image.Image(BytesIO(jpeg_bytes))
        img_obj.anchor = "B2"
        ws.add_image(img_obj)

        wb.save(file_path)

        # 提取图片
        extractor = MediaExtractor()
        media_objects = extractor.extract_from_xlsx(file_path)

        assert len(media_objects) == 1
        assert media_objects[0].raw_format.lower() in ("jpeg", "jpg", "image/jpeg")

    def test_extract_different_image_sizes(self, tmp_path: Path):
        """测试提取不同尺寸的图片"""
        file_path = tmp_path / "different_sizes.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active

        # 小图片
        small_img = Image.new("RGB", (10, 10), color="red")
        small_io = BytesIO()
        small_img.save(small_io, format="PNG")
        small_obj = openpyxl.drawing.image.Image(BytesIO(small_io.getvalue()))
        small_obj.anchor = "A1"
        ws.add_image(small_obj)

        # 大图片
        large_img = Image.new("RGB", (500, 500), color="blue")
        large_io = BytesIO()
        large_img.save(large_io, format="PNG")
        large_obj = openpyxl.drawing.image.Image(BytesIO(large_io.getvalue()))
        large_obj.anchor = "C3"
        ws.add_image(large_obj)

        wb.save(file_path)

        # 提取图片
        extractor = MediaExtractor()
        media_objects = extractor.extract_from_xlsx(file_path)

        assert len(media_objects) == 2
        # 验证图片尺寸不同
        assert len(media_objects[0].raw_data) != len(media_objects[1].raw_data)


class TestMediaExtractorIntegration:
    """测试 MediaExtractor 集成功能"""

    def test_extract_preserves_image_quality(self, xlsx_with_image: Path, sample_image: bytes):
        """测试提取过程保持图片质量"""
        extractor = MediaExtractor()
        media_objects = extractor.extract_from_xlsx(xlsx_with_image)

        assert len(media_objects) == 1

        # 验证提取的图片数据
        extracted_img = Image.open(BytesIO(media_objects[0].raw_data))
        original_img = Image.open(BytesIO(sample_image))

        # 验证尺寸相同
        assert extracted_img.size == original_img.size

        # 验证模式相同
        assert extracted_img.mode == original_img.mode
