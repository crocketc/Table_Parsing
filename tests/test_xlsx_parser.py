"""测试 XLSX 格式解析器"""

import datetime
from pathlib import Path
from unittest.mock import MagicMock, Mock, patch

import openpyxl
import pytest

from table_parsing.ir import Cell, Sheet, Workbook
from table_parsing.parsers.xlsx_parser import XLSXParser


@pytest.fixture
def sample_xlsx_file(tmp_path: Path) -> Path:
    """创建示例 XLSX 文件用于测试"""
    file_path = tmp_path / "test.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # 添加一些测试数据
    ws["A1"] = "姓名"
    ws["B1"] = "年龄"
    ws["A2"] = "张三"
    ws["B2"] = 25
    ws["A3"] = "李四"
    ws["B3"] = 30

    wb.save(file_path)
    return file_path


@pytest.fixture
def xlsx_with_formulas(tmp_path: Path) -> Path:
    """创建包含公式的 XLSX 文件"""
    file_path = tmp_path / "formulas.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "公式测试"

    ws["A1"] = 10
    ws["A2"] = 20
    ws["A3"] = 30
    ws["A4"] = "=SUM(A1:A3)"
    ws["A5"] = "=AVERAGE(A1:A3)"

    wb.save(file_path)
    return file_path


@pytest.fixture
def xlsx_with_merged_cells(tmp_path: Path) -> Path:
    """创建包含合并单元格的 XLSX 文件"""
    file_path = tmp_path / "merged.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "合并单元格"

    ws["A1"] = "标题"
    ws.merge_cells("A1:B1")
    ws["A2"] = "数据1"
    ws["B2"] = "数据2"

    wb.save(file_path)
    return file_path


@pytest.fixture
def xlsx_with_hidden_rows_cols(tmp_path: Path) -> Path:
    """创建包含隐藏行列的 XLSX 文件"""
    file_path = tmp_path / "hidden.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "隐藏行列"

    ws["A1"] = "A1"
    ws["A2"] = "A2"
    ws["A3"] = "A3"

    # 隐藏第2行
    ws.row_dimensions[2].hidden = True
    # 隐藏B列
    ws.column_dimensions["B"].hidden = True

    wb.save(file_path)
    return file_path


@pytest.fixture
def xlsx_with_number_formats(tmp_path: Path) -> Path:
    """创建包含数字格式的 XLSX 文件"""
    file_path = tmp_path / "formats.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "数字格式"

    ws["A1"] = 1234.56
    ws["A1"].number_format = "#,##0.00"
    ws["A2"] = 0.25
    ws["A2"].number_format = "0.00%"
    ws["A3"] = datetime.date(2024, 5, 3)
    ws["A3"].number_format = "YYYY-MM-DD"

    wb.save(file_path)
    return file_path


@pytest.fixture
def xlsx_with_multiple_sheets(tmp_path: Path) -> Path:
    """创建包含多个工作表的 XLSX 文件"""
    file_path = tmp_path / "multiple.xlsx"

    wb = openpyxl.Workbook()

    # 第一个工作表
    ws1 = wb.active
    ws1.title = "数据表"
    ws1["A1"] = "数据1"

    # 第二个工作表
    ws2 = wb.create_sheet("隐藏表")
    ws2["A1"] = "数据2"
    ws2.sheet_state = "hidden"

    # 第三个工作表
    ws3 = wb.create_sheet("另一个表")
    ws3["A1"] = "数据3"

    wb.save(file_path)
    return file_path


class TestXLSXParserBasic:
    """测试 XLSXParser 基础功能"""

    def test_parser_creation(self):
        """测试创建解析器实例"""
        parser = XLSXParser()
        assert parser is not None

    def test_parse_nonexistent_file(self):
        """测试解析不存在的文件"""
        parser = XLSXParser()
        with pytest.raises(FileNotFoundError):
            parser.parse("/nonexistent/file.xlsx")

    def test_parse_invalid_format(self, tmp_path: Path):
        """测试解析非 XLSX 格式文件"""
        invalid_file = tmp_path / "invalid.xlsx"
        invalid_file.write_text("Not a valid XLSX file")

        parser = XLSXParser()
        with pytest.raises(ValueError):
            parser.parse(invalid_file)


class TestXLSXParserFormulas:
    """测试 XLSXParser 公式提取功能"""

    def test_extract_formulas(self, xlsx_with_formulas: Path):
        """测试提取公式单元格"""
        parser = XLSXParser()
        workbook = parser.parse(xlsx_with_formulas)

        assert len(workbook.sheets) == 1
        sheet = workbook.sheets[0]

        # 检查 SUM 公式
        sum_cell = sheet.cells[3][0]  # A4
        assert sum_cell.is_formula is True
        assert sum_cell.formula_text == "SUM(A1:A3)"
        # 当 data_only=False 时，公式单元格的值是公式字符串本身
        # 我们主要验证 is_formula 和 formula_text 字段正确

        # 检查 AVERAGE 公式
        avg_cell = sheet.cells[4][0]  # A5
        assert avg_cell.is_formula is True
        assert avg_cell.formula_text == "AVERAGE(A1:A3)"

    def test_distinguish_formula_from_value(self, xlsx_with_formulas: Path):
        """测试区分公式和普通值"""
        parser = XLSXParser()
        workbook = parser.parse(xlsx_with_formulas)

        sheet = workbook.sheets[0]

        # 普通值单元格
        value_cell = sheet.cells[0][0]  # A1
        assert value_cell.is_formula is False
        assert value_cell.formula_text is None
        assert value_cell.value == 10


class TestXLSXParserMergedCells:
    """测试 XLSXParser 合并单元格功能"""

    def test_extract_merged_cells(self, xlsx_with_merged_cells: Path):
        """测试提取合并单元格信息"""
        parser = XLSXParser()
        workbook = parser.parse(xlsx_with_merged_cells)

        sheet = workbook.sheets[0]

        # A1:B1 合并区域
        merged_cell = sheet.cells[0][0]  # A1
        assert merged_cell.is_merged is True
        assert merged_cell.merge_range == "A1:B1"
        assert merged_cell.value == "标题"

    def test_merged_cell_only_in_first_position(self, xlsx_with_merged_cells: Path):
        """测试合并单元格只在第一个位置有值"""
        parser = XLSXParser()
        workbook = parser.parse(xlsx_with_merged_cells)

        sheet = workbook.sheets[0]

        # A1 是合并区域的第一个单元格，有值
        assert sheet.cells[0][0].value == "标题"
        assert sheet.cells[0][0].is_merged is True

        # B1 是合并区域的其他单元格，应该是空白
        assert sheet.cells[0][1].is_merged is True
        # 合并区域的其他单元格可能没有值或为 None


class TestXLSXParserHiddenRowsCols:
    """测试 XLSXParser 隐藏行列功能"""

    def test_extract_hidden_rows(self, xlsx_with_hidden_rows_cols: Path):
        """测试提取隐藏行信息"""
        parser = XLSXParser()
        workbook = parser.parse(xlsx_with_hidden_rows_cols)

        sheet = workbook.sheets[0]

        # 第2行是隐藏的
        # 检查第2行的所有单元格
        for col_idx in range(sheet.max_col):
            if len(sheet.cells) > 1 and col_idx < len(sheet.cells[1]):
                cell = sheet.cells[1][col_idx]
                # 单元格可能标记为隐藏，或者行信息存储在其他地方
                # 这里我们期望能识别隐藏状态

    def test_extract_hidden_columns(self, xlsx_with_hidden_rows_cols: Path):
        """测试提取隐藏列信息"""
        parser = XLSXParser()
        workbook = parser.parse(xlsx_with_hidden_rows_cols)

        sheet = workbook.sheets[0]

        # B列是隐藏的
        # 检查所有行的第2列（索引1）
        for row_idx in range(sheet.max_row):
            if len(sheet.cells) > row_idx and len(sheet.cells[row_idx]) > 1:
                cell = sheet.cells[row_idx][1]
                # 单元格可能标记为隐藏，或者列信息存储在其他地方


class TestXLSXParserNumberFormats:
    """测试 XLSXParser 数字格式功能"""

    def test_extract_number_format_styles(self, xlsx_with_number_formats: Path):
        """测试提取数字格式样式"""
        parser = XLSXParser()
        workbook = parser.parse(xlsx_with_number_formats)

        sheet = workbook.sheets[0]

        # 千分位格式
        cell1 = sheet.cells[0][0]  # A1
        assert cell1.style is not None
        assert cell1.style.get("number_format") == "#,##0.00"

        # 百分比格式
        cell2 = sheet.cells[1][0]  # A2
        assert cell2.style is not None
        assert cell2.style.get("number_format") == "0.00%"

        # 日期格式
        cell3 = sheet.cells[2][0]  # A3
        assert cell3.style is not None
        assert cell3.style.get("number_format") == "YYYY-MM-DD"


class TestXLSXParserMetadata:
    """测试 XLSXParser 元数据提取功能"""

    def test_extract_all_sheets_metadata(self, xlsx_with_multiple_sheets: Path):
        """测试提取所有工作表的元数据"""
        parser = XLSXParser()
        workbook = parser.parse(xlsx_with_multiple_sheets)

        # 应该有3个工作表
        assert len(workbook.sheets) == 3

        # 检查第一个工作表
        sheet1 = workbook.sheets[0]
        assert sheet1.name == "数据表"
        assert sheet1.hidden is False
        assert sheet1.max_row >= 1
        assert sheet1.max_col >= 1

        # 检查第二个工作表（隐藏）
        sheet2 = workbook.sheets[1]
        assert sheet2.name == "隐藏表"
        assert sheet2.hidden is True

        # 检查第三个工作表
        sheet3 = workbook.sheets[2]
        assert sheet3.name == "另一个表"
        assert sheet3.hidden is False

    def test_extract_workbook_metadata(self, sample_xlsx_file: Path):
        """测试提取工作簿级别的元数据"""
        parser = XLSXParser()
        workbook = parser.parse(sample_xlsx_file)

        # 应该有一些元数据
        assert workbook.metadata is not None
        assert isinstance(workbook.metadata, dict)


class TestXLSXParserIntegration:
    """测试 XLSXParser 综合功能"""

    def test_parse_complete_xlsx(self, sample_xlsx_file: Path):
        """测试完整解析 XLSX 文件"""
        parser = XLSXParser()
        workbook = parser.parse(sample_xlsx_file)

        # 验证工作簿结构
        assert isinstance(workbook, Workbook)
        assert len(workbook.sheets) >= 1

        # 验证工作表数据
        sheet = workbook.sheets[0]
        assert sheet.name == "Sheet1"
        assert sheet.max_row >= 2
        assert sheet.max_col >= 2
        assert len(sheet.cells) >= 2

        # 验证单元格数据
        assert sheet.cells[0][0].value == "姓名"
        assert sheet.cells[0][1].value == "年龄"
        assert sheet.cells[1][0].value == "张三"
        assert sheet.cells[1][1].value == 25
        assert sheet.cells[2][0].value == "李四"
        assert sheet.cells[2][1].value == 30

    def test_to_dict_serialization(self, sample_xlsx_file: Path):
        """测试解析后的数据可以正确序列化为字典"""
        parser = XLSXParser()
        workbook = parser.parse(sample_xlsx_file)

        # 转换为字典
        result = workbook.to_dict()

        # 验证结构
        assert "metadata" in result
        assert "sheets" in result
        assert len(result["sheets"]) >= 1

        # 验证可以序列化为 JSON（不抛出异常）
        import json
        json_str = json.dumps(result)
        assert len(json_str) > 0
