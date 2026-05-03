"""
XLSX 格式解析器
"""

import datetime
from pathlib import Path
from typing import Union

import openpyxl
from openpyxl.utils import get_column_letter

from ..ir.model import Cell, Sheet, Workbook
from .base import BaseParser


class XLSXParser(BaseParser):
    """
    XLSX 格式解析器

    解析 Excel 2007+ (XLSX) 文件并转换为统一的中间表示

    Features:
    - 解析所有工作表及其元数据
    - 提取公式信息（is_formula, formula_text, cached value）
    - 提取合并单元格区域
    - 提取隐藏的行和列
    - 提取数字格式样式
    """

    def parse(self, file_path: Union[str, Path]) -> Workbook:
        """
        解析 XLSX 文件

        Args:
            file_path: XLSX 文件路径

        Returns:
            Workbook 对象，包含解析后的数据

        Raises:
            FileNotFoundError: 文件不存在
            ValueError: 文件格式不正确
        """
        path = Path(file_path)

        # 检查文件是否存在
        if not path.exists():
            raise FileNotFoundError(f"XLSX file not found: {file_path}")

        # 检查文件扩展名
        if path.suffix.lower() not in (".xlsx", ".xlsm"):
            raise ValueError(
                f"Invalid file extension: {path.suffix}. Expected .xlsx or .xlsm"
            )

        try:
            # 使用 openpyxl 加载工作簿
            # data_only=False 确保能读取公式本身，而不仅仅是计算结果
            wb = openpyxl.load_workbook(path, data_only=False, keep_vba=False)

            # 提取工作簿元数据
            metadata = self._extract_workbook_metadata(wb, path)

            # 解析所有工作表
            sheets = []
            for sheet in wb.worksheets:
                parsed_sheet = self._parse_sheet(sheet)
                sheets.append(parsed_sheet)

            wb.close()

            return Workbook(metadata=metadata, sheets=sheets)

        except Exception as e:
            if isinstance(e, (FileNotFoundError, ValueError)):
                raise
            raise ValueError(f"Failed to parse XLSX file: {e}") from e

    def _extract_workbook_metadata(self, wb: openpyxl.Workbook, path: Path) -> dict:
        """
        提取工作簿级别的元数据

        Args:
            wb: openpyxl Workbook 对象
            path: 文件路径

        Returns:
            元数据字典
        """
        metadata = {
            "filename": path.name,
        }

        # 尝试获取更多元数据（如果存在）
        if hasattr(wb, "properties") and wb.properties:
            props = wb.properties
            if props.creator:
                metadata["creator"] = props.creator
            if props.created:
                metadata["created"] = props.created.isoformat()
            if props.modified:
                metadata["modified"] = props.modified.isoformat()
            if props.title:
                metadata["title"] = props.title
            if props.subject:
                metadata["subject"] = props.subject

        return metadata

    def _parse_sheet(self, sheet: openpyxl.worksheet.worksheet.Worksheet) -> Sheet:
        """
        解析单个工作表

        Args:
            sheet: openpyxl Worksheet 对象

        Returns:
            Sheet 对象
        """
        # 获取工作表基本信息
        name = sheet.title
        hidden = sheet.sheet_state == "hidden"

        # 获取工作表的实际使用范围
        max_row = sheet.max_row if sheet.max_row else 0
        max_col = sheet.max_column if sheet.max_column else 0

        # 提取合并单元格信息
        merged_ranges = list(sheet.merged_cells.ranges)
        merged_cells_map = self._build_merged_cells_map(merged_ranges)

        # 提取隐藏行和列
        hidden_rows = self._extract_hidden_rows(sheet)
        hidden_cols = self._extract_hidden_columns(sheet)

        # 解析所有单元格
        cells = []
        for row_idx in range(1, max_row + 1):
            row_cells = []
            for col_idx in range(1, max_col + 1):
                cell = self._parse_cell(
                    sheet,
                    row_idx,
                    col_idx,
                    merged_cells_map,
                    hidden_rows,
                    hidden_cols,
                )
                row_cells.append(cell)
            cells.append(row_cells)

        return Sheet(
            name=name,
            hidden=hidden,
            max_row=max_row,
            max_col=max_col,
            cells=cells,
        )

    def _build_merged_cells_map(
        self, merged_ranges: list
    ) -> dict[tuple[int, int], str]:
        """
        构建合并单元格映射表

        Args:
            merged_ranges: 合并单元格区域列表

        Returns:
            字典，键为 (row, column)，值为合并区域字符串（如 "A1:B2"）
        """
        merged_map = {}
        for merged_range in merged_ranges:
            # 获取合并区域的起始和结束坐标
            min_col, min_row, max_col, max_row = merged_range.bounds

            # 为合并区域内的所有单元格创建映射
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    # 将合并区域转换为 Excel 坐标表示
                    range_str = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
                    merged_map[(row, col)] = range_str

        return merged_map

    def _extract_hidden_rows(
        self, sheet: openpyxl.worksheet.worksheet.Worksheet
    ) -> set[int]:
        """
        提取隐藏行的索引

        Args:
            sheet: openpyxl Worksheet 对象

        Returns:
            隐藏行索引的集合
        """
        hidden_rows = set()
        for row_idx, row_dimension in sheet.row_dimensions.items():
            if row_dimension.hidden:
                hidden_rows.add(row_idx)
        return hidden_rows

    def _extract_hidden_columns(
        self, sheet: openpyxl.worksheet.worksheet.Worksheet
    ) -> set[int]:
        """
        提取隐藏列的索引

        Args:
            sheet: openpyxl Worksheet 对象

        Returns:
            隐藏列索引的集合
        """
        hidden_cols = set()
        for col_letter, col_dimension in sheet.column_dimensions.items():
            if col_dimension.hidden:
                # 将列字母转换为数字索引
                col_idx = openpyxl.utils.column_index_from_string(col_letter)
                hidden_cols.add(col_idx)
        return hidden_cols

    def _parse_cell(
        self,
        sheet: openpyxl.worksheet.worksheet.Worksheet,
        row_idx: int,
        col_idx: int,
        merged_cells_map: dict[tuple[int, int], str],
        hidden_rows: set[int],
        hidden_cols: set[int],
    ) -> Cell:
        """
        解析单个单元格

        Args:
            sheet: openpyxl Worksheet 对象
            row_idx: 行索引（从1开始）
            col_idx: 列索引（从1开始）
            merged_cells_map: 合并单元格映射表
            hidden_rows: 隐藏行索引集合
            hidden_cols: 隐藏列索引集合

        Returns:
            Cell 对象
        """
        # 获取 openpyxl 单元格对象
        openpyxl_cell = sheet.cell(row_idx, col_idx)

        # 确定单元格值和数据类型
        value, data_type, raw_value = self._extract_cell_value(openpyxl_cell)

        # 检查是否是公式
        is_formula = openpyxl_cell.data_type == "f"  # 'f' 表示公式
        formula_text = None
        if is_formula and openpyxl_cell.value:
            # openpyxl 中公式以 = 开头
            formula_text = str(openpyxl_cell.value).lstrip("=")

        # 检查是否是合并单元格
        is_merged = (row_idx, col_idx) in merged_cells_map
        merge_range = merged_cells_map.get((row_idx, col_idx)) if is_merged else None

        # 检查是否隐藏
        is_hidden = (row_idx in hidden_rows) or (col_idx in hidden_cols)

        # 提取样式信息（数字格式等）
        style = self._extract_cell_style(openpyxl_cell)

        return Cell(
            value=value,
            raw_value=raw_value,
            data_type=data_type,
            is_formula=is_formula,
            formula_text=formula_text,
            is_merged=is_merged,
            merge_range=merge_range,
            is_hidden=is_hidden,
            style=style,
        )

    def _extract_cell_value(
        self, cell: openpyxl.cell.cell.Cell
    ) -> tuple:
        """
        提取单元格的值、数据类型和原始值

        Args:
            cell: openpyxl Cell 对象

        Returns:
            (value, data_type, raw_value) 元组
        """
        raw_value = None
        value = None
        data_type = "string"

        if cell.value is None:
            # 空单元格
            value = None
            data_type = "blank"
        elif cell.data_type == "f":
            # 公式单元格
            # 对于公式，我们使用计算后的值作为 value
            # 但 formula_text 会在另一个字段中存储
            try:
                # 尝试获取公式的计算结果
                value = cell.value
                if isinstance(value, (int, float)):
                    data_type = "number"
                elif isinstance(value, bool):
                    data_type = "bool"
                elif isinstance(value, (datetime.datetime, datetime.date)):
                    data_type = "date"
                else:
                    data_type = "string"
            except Exception:
                value = None
                data_type = "blank"
        elif cell.data_type == "n":
            # 数字类型
            value = cell.value
            raw_value = str(value)
            data_type = "number"
        elif cell.data_type == "s":
            # 字符串类型
            value = str(cell.value)
            raw_value = value
            data_type = "string"
        elif cell.data_type == "d":
            # 日期类型
            value = cell.value
            raw_value = str(value)
            data_type = "date"
        elif cell.data_type == "b":
            # 布尔类型
            value = bool(cell.value)
            raw_value = str(value)
            data_type = "bool"
        else:
            # 其他类型，作为字符串处理
            value = str(cell.value) if cell.value is not None else None
            raw_value = value
            data_type = "string" if value is not None else "blank"

        return (value, data_type, raw_value)

    def _extract_cell_style(self, cell: openpyxl.cell.cell.Cell) -> dict | None:
        """
        提取单元格样式信息

        Args:
            cell: openpyxl Cell 对象

        Returns:
            样式字典，如果没有样式则为 None
        """
        style = {}

        # 提取数字格式
        if cell.number_format and cell.number_format != "General":
            style["number_format"] = cell.number_format

        # 可以在这里添加更多样式提取逻辑
        # 例如：字体、对齐方式、边框、填充等

        return style if style else None
